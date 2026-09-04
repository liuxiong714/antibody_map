"""Submodule of app.services.analysis (split from analysis_service.py).

导出层：将各分析结果构建为 Excel（多 sheet）字节流，供 API 层下载；并 re-export
数据集快照数据函数（get_approved_data_points_for_snapshot）。"""

import io

# 数据集快照数据由 get_approved_data_points_for_snapshot（data_management）提供，此处 re-export 以保持结构完整。
from app.services.analysis.data_management import (
    get_approved_data_points_for_snapshot,  # noqa: F401
)


def build_excel_export(
    trend_data,
    region_data,
    age_data,
    summary_data,
    approved_items,
) -> bytes:
    """将所有分析结果构建为 Excel 多 sheet，返回文件字节流。

    生成 sheet：汇总统计、年份趋势、区域对比、年龄分层、数据点明细、统计方法附录
    （加权率/GMC/95%CI/基尼/meta合并等算法公式）。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _write_rows(ws, rows, start_row=2):
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="center")

    # Sheet 1: 汇总统计
    ws_summary = wb.active
    ws_summary.title = "汇总统计"
    _write_header(ws_summary, ["指标", "数值"])
    s = summary_data if isinstance(summary_data, dict) else {}
    _write_rows(ws_summary, [[k, v] for k, v in s.items()])
    for col in range(1, 3):
        ws_summary.column_dimensions[chr(64 + col)].width = 25

    # Sheet 2: 年份趋势（get_trend 返回 {"trend": [...], "trend_significance": {...}}）
    trend_rows = trend_data.get("trend", []) if isinstance(trend_data, dict) else (trend_data or [])
    if trend_rows:
        ws_trend = wb.create_sheet("年份趋势")
        trend_keys = list(trend_rows[0].keys()) if isinstance(trend_rows[0], dict) else []
        _write_header(ws_trend, trend_keys)
        for i, row in enumerate(trend_rows, 2):
            if isinstance(row, dict):
                for j, k in enumerate(trend_keys, 1):
                    ws_trend.cell(row=i, column=j, value=row.get(k))
        for col_idx in range(1, len(trend_keys) + 1):
            ws_trend.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18

    # Sheet 3: 区域对比
    if region_data:
        ws_region = wb.create_sheet("区域对比")
        region_keys = list(region_data[0].keys()) if isinstance(region_data[0], dict) else []
        _write_header(ws_region, region_keys)
        for i, row in enumerate(region_data, 2):
            if isinstance(row, dict):
                for j, k in enumerate(region_keys, 1):
                    ws_region.cell(row=i, column=j, value=row.get(k))
        for col_idx in range(1, len(region_keys) + 1):
            ws_region.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18

    # Sheet 4: 年龄分层
    if age_data:
        ws_age = wb.create_sheet("年龄分层")
        age_keys = list(age_data[0].keys()) if isinstance(age_data[0], dict) else []
        _write_header(ws_age, age_keys)
        for i, row in enumerate(age_data, 2):
            if isinstance(row, dict):
                for j, k in enumerate(age_keys, 1):
                    ws_age.cell(row=i, column=j, value=row.get(k))
        for col_idx in range(1, len(age_keys) + 1):
            ws_age.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18

    # Sheet 5: 数据点明细
    if approved_items:
        ws_dp = wb.create_sheet("数据点明细")
        dp_fields = [
            "literature_title", "disease", "province", "city", "age_group",
            "sample_size", "data_type", "value", "unit", "ci_lower", "ci_upper",
            "collection_year", "method", "population", "confidence",
        ]
        dp_headers = [
            "文献标题", "疾病", "省份", "城市", "年龄组", "样本量", "数据类型",
            "数值", "单位", "CI下限", "CI上限", "采集年份", "检测方法", "人群", "置信度",
        ]
        # P1-3：导出新增作者/DOI/PMID/引用列
        dp_extra_fields = ["literature_authors", "literature_doi", "literature_pmid", "literature_citation"]
        dp_extra_headers = ["文献作者", "DOI", "PMID", "引用"]
        dp_fields = dp_fields + dp_extra_fields
        dp_headers = dp_headers + dp_extra_headers
        _write_header(ws_dp, dp_headers)
        for i, item in enumerate(approved_items, 2):
            if isinstance(item, dict):
                for j, field in enumerate(dp_fields, 1):
                    val = item.get(field)
                    if field == "literature_citation":
                        # 构建引用字符串：作者. 标题. 期刊. 年份. doi: DOI
                        _parts = []
                        _a = (item.get("literature_authors") or "").strip()
                        _t = (item.get("literature_title") or "").strip()
                        _jn = (item.get("literature_journal") or "").strip()
                        _yr = item.get("literature_year")
                        _doi = (item.get("literature_doi") or "").strip()
                        if _a:
                            _parts.append(_a)
                        if _t:
                            _parts.append(_t)
                        if _jn:
                            _parts.append(_jn)
                        if _yr:
                            _parts.append(str(_yr))
                        if _doi:
                            _parts.append(f"doi: {_doi}")
                        val = ". ".join(_parts)
                    ws_dp.cell(row=i, column=j, value=val)
        for col_idx in range(1, len(dp_headers) + 1):
            ws_dp.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18

    # Sheet 6: 统计方法附录（算法与公式说明）
    ws_methods = wb.create_sheet("统计方法附录")
    _write_header(ws_methods, ["统计方法", "公式", "说明"])
    wrap_align = Alignment(vertical="top", wrap_text=True)
    appendix_rows = [
        (
            "加权阳性率（逆方差加权合并）",
            "w_i = n_i / (p_i·(1-p_i))\npooled = Σ(w_i·p_i) / Σw_i",
            "p_i = 研究阳性率（0-1，>1 视为百分数除以 100）；n_i = 样本量。"
            "p=0 或 1 的边界研究采用连续性校正（x'=x+0.5，n'=n+1）避免方差为 0 / 权重无穷大。"
            "输出为百分比（×100）。",
        ),
        (
            "95% 置信区间（Wilson score）",
            "center = (p + z²/2n) / (1 + z²/n)\n"
            "half = z·√(p(1-p)/n + z²/4n²) / (1 + z²/n)\n"
            "CI = [max(0, center-half), min(1, center+half)]",
            "z = 1.96（95% 置信水平）；n 取合并总样本量。用于合并阳性率的区间估计。",
        ),
        (
            "GMC 几何均数 + 对数域 t 分布 CI",
            "GMC = exp( mean(ln v_i) )\nSE = SD(ln v) / √n\n"
            "CI = exp( mean(ln v) ± t(n-1)·SE )",
            "仅对正数计算，缺失 / 非正值剔除。t 临界值查表（df>30 用 z=1.96 近似）。"
            "几何均数适用于抗体滴度等对数正态分布数据。",
        ),
        (
            "基尼系数（省间公平性）",
            "G = 2·Σ(i+1)·x_i / (n·Σx_i) - (n+1)/n",
            "对排序后的非负序列计算（i 从 0 起）。G=0 完全均等，G=1 完全不均等。"
            "用于评估各省加权阳性率的离散程度。",
        ),
        (
            "变异系数",
            "CV = 样本标准差 / |均值|",
            "无量纲比值，衡量省际阳性率相对离散程度。样本数 < 2 或均值为 0 时无法计算，返回 0。",
        ),
        (
            "Meta 合并（逆方差 固定/随机效应 + I²）",
            "var_i = ((CI_upper - CI_lower) / (2·1.96))²\n"
            "固定: pooled = Σ(w_i·p_i)/Σw_i，w_i = 1/var_i\n"
            "Q = Σ w_i·(p_i - pooled)²，df = k-1\n"
            "I² = max(0, (Q - df)/Q)·100%\n"
            "随机(D-L): τ² = max(0,(Q-df)/(Σw-Σw²/Σw))，w*_i = 1/(var_i+τ²)",
            "CI 缺失时退化为二项方差 p(1-p)/n（p=0/1 连续性校正）。"
            "异质性解读：I²<25% 低、25%-50% 中、>50% 高。",
        ),
        (
            "惩罚样条平滑（年龄曲线）+ FOI",
            "logit(P) = Bβ（自然三次 B-spline，k=min(8,max(3,n//3))）\n"
            "NLL = Σ n·[y·log p + (1−y)·log(1−p)] + λp·Σ(Δ²β)²\n"
            "GCV 选 λp∈{1e-3..1e2 对数12点}\n"
            "FOI: λ(a) = P′(a)/(1−P(a))，P′ 用样条解析导数",
            "对已审核主估计按 age_mid 聚合（无则用 _midpoint_age 推算，仍无则剔除计数）后拟合；"
            "置信带用 delta 法 ±1.96·√(b·Cov·bᵀ) 逆 logit 还原，输出 0.5 岁步长曲线。"
            "数据点 <8 返回 422。FOI 数值安全：P≥0.999 时 λ 置 None。",
        ),
        (
            "加权线性趋势（显著性）",
            "y = a + b·x；b = Sxy/Sxx\nR² = 1 - SS_res/SS_tot；p 值 = 斜率 t 检验",
            "权重取各年总样本量。趋势方向：斜率 >0 上升、<0 下降、≈0 平稳。",
        ),
        (
            "证据可靠性分级 A/B/C/D",
            "score = 样本量分 + CI分 + 置信度分 + 溯源分 + 研究数分",
            "样本量 ≥1000:+4 / ≥300:+3 / ≥100:+2 / ≥30:+1；带CI:+2；置信度 high:+2 / medium:+1；"
            "原文溯源:+2；研究数 ≥5:+2 / ≥2:+1。总分 ≥9→A、≥6→B、≥3→C、<3→D。",
        ),
        (
            "FOI 催化模型（免疫屏障模拟）",
            "λ = -ln(1-SP) / age\nR0 ≈ λ·L（L=期望寿命 75 年）\nHIT = (1 - 1/R0)·100%",
            "SP=观测血清阳性率（0-1），age=年龄组中点。模拟有效免疫比例 "
            "effective = 覆盖 + (1-覆盖)·加强针比例，与 HIT 对比判定屏障状态并反推达标所需覆盖。",
        ),
    ]
    for r_idx, (method, formula, note) in enumerate(appendix_rows, 2):
        ws_methods.cell(row=r_idx, column=1, value=method).alignment = wrap_align
        ws_methods.cell(row=r_idx, column=2, value=formula).alignment = wrap_align
        ws_methods.cell(row=r_idx, column=3, value=note).alignment = wrap_align
        ws_methods.row_dimensions[r_idx].height = 80
    ws_methods.column_dimensions["A"].width = 28
    ws_methods.column_dimensions["B"].width = 46
    ws_methods.column_dimensions["C"].width = 60

    # 序列化到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
