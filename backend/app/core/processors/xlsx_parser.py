"""XLSX 解析器：openpyxl 读每个 sheet → 逐行渲染为 GFM Markdown 表格。"""
import io

from .base import BaseParser, grid_to_markdown, register_parser


@register_parser(".xlsx")
class XlsxParser(BaseParser):
    supported_ext = ".xlsx"

    def extract(self, file_bytes: bytes) -> str:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            md = grid_to_markdown(rows, f"### 工作表: {sheet_name}")
            if md:
                parts.append(md)
            else:
                # 空表：仅输出工作表标题，保留表的存在信息
                parts.append(f"### 工作表: {sheet_name}")
        wb.close()
        return "\n\n".join(parts)